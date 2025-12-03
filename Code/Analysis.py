import openpyxl
import csv
import os
import shutil

from openpyxl.chart import (
    BarChart,
    Reference,
)
from openpyxl.chart.series import SeriesLabel



def Load(csv_path):
    wb = openpyxl.Workbook()
    ws = wb.active

    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f, delimiter=',')

        for row in reader:
            ws.append(row)
    
    for row in ws.iter_rows():
        for x in row:
            try:
                x.value = float(x.value)
            except:
                pass
    return wb

def BarChartGen(wb, title, x_axis, y_axis):
    ws = wb.active
    starts = [2]
    series = []
    graph_type = [" ADF Male", " ADF Female", " ADF Persons", " Population"]
    y = 0
    for col in ws.iter_cols():
        if y > 1:
            series.append(col[0].value)
        y += 1
    for row in ws.iter_rows():
        if row[0].row > 2 and ws.cell(row=row[0].row, column=1).value != ws.cell(row=row[0].row - 1, column=1).value:
            starts.append(row[0].row)
    for i in range(0, 3):
        start_row = starts[i] if i < len(starts) else None
        if start_row is None:
            return wb
        end_row = starts[i+1] - 1 if i + 1 < len(starts) else ws.max_row
        c1 = BarChart()
        c1.title = (title  + graph_type[i])
        c1.style = 2
        c1.y_axis.title = y_axis
        c1.x_axis.title = x_axis
        data = Reference(ws, min_col=3, max_col=ws.max_column - 1, min_row=start_row, max_row=end_row)
        c1.add_data(data, titles_from_data=False)
        for i in range(len(series)):
           c1.series[i - 1].title = SeriesLabel(v=series[i - 1])
        cats = Reference(ws, min_col=2, max_col=2, min_row=2, max_row=ws.max_row)
        c1.set_categories(cats)
        c1.x_axis.delete = False
        c1.y_axis.delete = False
        ws.add_chart(c1, "G{}".format(start_row))
        if end_row == ws.max_row:
            c1 = BarChart()
            c1.title = (title + graph_type[len(graph_type) - 1])
            c1.style = 2
            c1.y_axis.title = y_axis
            c1.x_axis.title = x_axis
            data = Reference(ws, min_col=ws.max_column, max_col=ws.max_column, min_row=start_row, max_row=end_row)
            c1.add_data(data, titles_from_data=False)
            c1.series[0].title = SeriesLabel(v=series[len(series) - 1])
            cats = Reference(ws, min_col=2, max_col=2, min_row=2, max_row=ws.max_row)
            c1.set_categories(cats)
            c1.x_axis.delete = False
            c1.y_axis.delete = False
            ws.add_chart(c1, "P{}".format(start_row))
    return wb
    
    
def Analysis(input_dir, output_dir):
    os.makedirs(output_dir, exist_ok=True)

    # clone folder structure and analyse csv files into excels
    dirpath, dirnames, filenames = next(os.walk(input_dir))
    for dirpath, dirnames, filenames in os.walk(input_dir):
        rel = os.path.relpath(dirpath, input_dir)
        target_dir = os.path.join(output_dir, rel) if rel != "." else output_dir
        os.makedirs(target_dir, exist_ok=True)
        for name in filenames:
            if not name.lower().endswith(".csv"):
                continue

            src_csv = os.path.join(dirpath, name)
            dst_csv = os.path.join(target_dir, name)
            shutil.copy2(src_csv, dst_csv)

            # make excel file
            wb = Load(dst_csv)
            wb = BarChartGen(wb, name, "Category", "Count")

            xlsx_path = os.path.join(output_dir, os.path.splitext(dst_csv)[0] + ".xlsx")
            wb.save(xlsx_path)

            # delete analysed csv
            os.remove(dst_csv)
    return