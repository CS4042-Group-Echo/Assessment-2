import os
import shutil
import openpyxl
import csv


def IsDividerRow(Sheet, RowNumber):  
    MaxColumn = Sheet.max_column
    Values = []
    for ColumnNumber in range(1, MaxColumn + 1):
        Value = Sheet.cell(row=RowNumber, column=ColumnNumber).value
        if Value not in (None, ""):
            Values.append(str(Value).strip())
    if len(Values) != 1:
        return False
    Text = Values[0]
    return Text.isupper() and any(ch.isalpha() for ch in Text)


def SafeFolder(text):  
    if text is None:
        return "unknown"
    text = str(text).strip().lower()
    for ch in '<>:"/\\|?*':
        text = text.replace(ch, "_")
    text = "_".join(text.split())
    return text


def FormatFile(input_dir, formatted_dir):
    os.makedirs(formatted_dir, exist_ok=True)

    for filename in os.listdir(input_dir):
        if not filename.endswith(".xlsx"):
            continue

        Source = os.path.join(input_dir, filename)
        NewName = "A_" + filename

        # ADF / Populis layer
        if filename.lower().startswith("adf"):
            OwnerGroup = "ADF"
        else:
            OwnerGroup = "Populis"

        OwnerFolder = os.path.join(formatted_dir, OwnerGroup)
        os.makedirs(OwnerFolder, exist_ok=True)

        Destination = os.path.join(OwnerFolder, NewName)
        shutil.copy(Source, Destination)
        print(f"[FormatFile] Copied → {Destination}")

        Workbook = openpyxl.load_workbook(Destination)

        for Index, SheetName in enumerate(Workbook.sheetnames):
            if Index == 0:
                continue

            Sheet = Workbook[SheetName]

            # A6 becomes container folder for all split tables
            MetaA5 = Sheet["A5"].value
            MetaA6 = Sheet["A6"].value

            Part5 = SafeFolder(MetaA5 or "")
            Part6 = SafeFolder(MetaA6 or "")

            FolderName = (Part5 + "__" + Part6).strip("_")
            FolderPath = os.path.join(OwnerFolder, FolderName)   # *** FIXED PATH ***
            os.makedirs(FolderPath, exist_ok=True)

            Sheet.delete_rows(1, 7)

            for RowNumber in range(Sheet.max_row, 0, -1):
                if all(
                    Sheet.cell(row=RowNumber, column=ColumnNumber).value in (None, "")
                    for ColumnNumber in range(1, Sheet.max_column + 1)
                ):
                    Sheet.delete_rows(RowNumber)

            FirstDataRow = None
            for RowNumber in range(1, Sheet.max_row + 1):
                v = Sheet.cell(row=RowNumber, column=1).value
                if v in (None, ""):
                    continue
                if IsDividerRow(Sheet, RowNumber):
                    continue
                FirstDataRow = RowNumber
                break

            if FirstDataRow is None or FirstDataRow <= 1:
                continue

            HeaderRows = list(range(1, FirstDataRow))
            NonDividerHeaderRows = [r for r in HeaderRows if not IsDividerRow(Sheet, r)]
            MaxColumn = Sheet.max_column

            Combined = {}
            for c in range(1, MaxColumn + 1):
                parts = []
                for r in NonDividerHeaderRows:
                    v = Sheet.cell(row=r, column=c).value
                    if v not in (None, ""):
                        parts.append(str(v).strip())
                Combined[c] = " ".join(parts) if parts else None

            Sheet.insert_rows(1)

            for c in range(1, MaxColumn + 1):
                if Combined[c] not in (None, ""):
                    Sheet.cell(row=1, column=c, value=Combined[c])

            ToDelete = [r + 1 for r in NonDividerHeaderRows]
            for r in sorted(ToDelete, reverse=True):
                Sheet.delete_rows(r, 1)

            # Identify divider rows (start of each table)
            DividerRows = []
            for r in range(2, Sheet.max_row + 1):
                if IsDividerRow(Sheet, r):
                    DividerRows.append(r)

            # If NO dividers → save the whole sheet as all.csv
            if not DividerRows:
                CsvPath = os.path.join(FolderPath, "all.csv")

                with open(CsvPath, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)

                    # header first
                    header = [Sheet.cell(row=1, column=c).value for c in range(1, MaxColumn + 1)]
                    writer.writerow(header)

                    # every data row
                    for r in range(2, Sheet.max_row + 1):
                        row = [Sheet.cell(row=r, column=c).value for c in range(1, MaxColumn + 1)]
                        writer.writerow(row)

                print(f"[FormatFile] CSV (no dividers → all.csv) → {CsvPath}")
                continue

            for index, div_row in enumerate(DividerRows):
                Start = div_row
                End = DividerRows[index + 1] - 1 if index + 1 < len(DividerRows) else Sheet.max_row

                divider_cell = [
                    Sheet.cell(row=div_row, column=c).value
                    for c in range(1, MaxColumn + 1)
                    if Sheet.cell(row=div_row, column=c).value not in (None, "")
                ]
                TableName = divider_cell[0].strip() if divider_cell else f"table_{index+1}"
                TableNameSafe = TableName.lower().replace(" ", "_")

                CsvPath = os.path.join(FolderPath, f"{TableNameSafe}.csv")
                with open(CsvPath, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    header = [Sheet.cell(row=1, column=c).value for c in range(1, MaxColumn + 1)]
                    writer.writerow(header)
                    for r in range(Start, End + 1):
                        row = [Sheet.cell(row=r, column=c).value for c in range(1, MaxColumn + 1)]
                        writer.writerow(row)

                print(f"[FormatFile] CSV → {CsvPath}")
        # After CSV extraction, delete the temporary workbook
        os.remove(Destination)
        print(f"[FormatFile] Done: {Destination}\n")
