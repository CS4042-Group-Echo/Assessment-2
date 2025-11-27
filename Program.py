import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
# import numpy as np
# import matplotlib.pyplot as plt

file_path_ADF = "InputData\AdfInputData.xlsx"
file_path_Pop = "InputData\PopulusInputData.xlsx"
file_path_Clean_ADF = "CleanData\AdfCleanData.xlsx"
file_path_Clean_Pop = "CleanData\PopulusCleanData.xlsx"
file_path_Output = "OutputData\OutputData.xlsx"


def LineChartGen(wbPath, wsName, title, x_title, y_title, data_range, category_range, pos):
    wb = load_workbook(wbPath)
    ws = wb[wsName]
    c1 = LineChart()
    c1.title = title
    c1.style = 13
    c1.y_axis.title = y_title
    c1.x_axis.title = x_title

    data = Reference(ws, min_col=data_range[0], min_row=data_range[1], max_col=data_range[2], max_row=data_range[3])
    c1.add_data(data)
    cats = Reference(ws, min_col=category_range[0], min_row=category_range[1], max_col=category_range[0], max_row=category_range[2])
    c1.set_categories(cats)
    # Style the lines
    LineCount = len(c1.series)
    i = 1
    for series in c1.series:
        rVal = (256/LineCount) * i
        gVal = 128
        bVal = 256 - (256/LineCount) * i
        hexColor = "{:02X}{:02X}{:02X}".format(int(rVal), int(gVal), int(bVal))
        series.graphicalProperties.line.solidFill = hexColor
        i += 1
    wb = openpyxl.Workbook()
    ws.add_chart(c1, 'A1')
    wb.save(file_path_Output)


#placeholder for data cleaning process
wb = openpyxl.load_workbook(file_path_ADF)
wb.save(file_path_Clean_ADF)
wb = openpyxl.load_workbook(file_path_Pop)
wb.save(file_path_Clean_Pop)


wb = load_workbook('CleanData\AdfCleanData.xlsx')
ws = wb.active
LineChartGen('CleanData\AdfCleanData.xlsx', ws.title, 'Age Distribution', 'Age Groups', 'Percentage', (2, 13, 4, 27), (1, 13, 27), 'C' + str(ws.max_row + 2))